import streamlit as st
import pandas as pd
import os
from pathlib import Path
import email
from email.policy import default
import xlwings as xw

# Función para guardar adjuntos de correos
def guardar_adjuntos_eml(uploaded_files, carpeta_salida):
    Path(carpeta_salida).mkdir(parents=True, exist_ok=True)
    for uploaded_file in uploaded_files:
        mensaje = email.message_from_bytes(uploaded_file.read(), policy=default)
        for parte in mensaje.walk():
            if parte.get_content_disposition() == "attachment":
                filename = parte.get_filename()
                if filename and (filename.endswith(".xlsx") or filename.endswith(".xls")):
                    ruta_salida = os.path.join(carpeta_salida, filename)
                    with open(ruta_salida, "wb") as archivo_salida:
                        archivo_salida.write(parte.get_payload(decode=True))

# Función para consolidar archivos
def consolidar_archivos(directorio):
    dataframes_validos = []
    for archivo in os.listdir(directorio):
        if archivo.endswith(".xlsx") or archivo.endswith(".xls"):
            ruta_archivo = os.path.join(directorio, archivo)
            xls = pd.ExcelFile(ruta_archivo)
            for hoja in xls.sheet_names:
                df = pd.read_excel(xls, hoja)
                if "CICLO INSCRIPCIÓN" in df.columns:
                    df.rename(columns={"CICLO INSCRIPCIÓN": "CICLO"}, inplace=True)
                if df.shape[1] == 25:
                    dataframes_validos.append(df)
    df_unificado = pd.concat(dataframes_validos, ignore_index=True)
    df_unificado['TIPO'] = df_unificado['NOMBRE DE EMPRESA  PROPULSOR  CONEMPLEO'].apply(
        lambda x: 'Empresa' if pd.notna(x) and x.strip() != '' else 'Cesante'
    )
    # Reemplazos en la columna "TALLER Y/O CURSO"
    reemplazos = {
        "DESARROLLO DE INNOVACIONES INTERNAS  INTRAEMPRENDIMIENTO": "DESARROLLO DE INNOVACIONES INTERNAS – INTRAEMPRENDIMIENTO",
        "ECOMMERCE Y NEGOCIOS DIGITALES": "E-COMMERCE Y NEGOCIOS DIGITALES",
        "EXCEL BASICO INTERMEDIO": "EXCEL BASICO - INTERMEDIO",
        "EXCEL BASICO INTERMEDIO ": "EXCEL BASICO - INTERMEDIO",
        "EXCEL INTERMEDIO AVANZADO": "EXCEL INTERMEDIO - AVANZADO",
        "CONSTRUCCION DE INDICADORES ":"CONSTRUCCION DE INDICADORES",
        "LENGUA DE SEÑAS PARA EL SERVICIO ": "LENGUA DE SEÑAS PARA EL SERVICIO",
        "LENGUAJE DE VENTAS ": "LENGUAJE DE VENTAS",
        "POWER BI PARA GESTION ADMINISTRATIVA": "POWER BI PARA LA GESTION ADMINISTRATIVA",
        "POWER BI PARA LA GESTION ADMINISTRATIVA ": "POWER BI PARA LA GESTION ADMINISTRATIVA",
        "REDACCION Y ORTOGRAFIA TECNICAS Y HABILIADADES PARA LA COMUNICACION ESCRITA": "REDACCION Y ORTOGRAFIA TECNICAS Y HABILIDADES PARA COMUNICACION ESCRITA",
        "REDACCION Y ORTOGRAFIA, TECNICAS Y HABILIDADES PARA COMUNICACION ESCRITA": "REDACCION Y ORTOGRAFIA TECNICAS Y HABILIDADES PARA COMUNICACION ESCRITA",
        "VOCACION DE SERVICIO AL CLIENTE ": "VOCACION DE SERVICIO AL CLIENTE",
        "CONVERSATIONAL ENGLISH FOR BPO'S":"CONVERSATIONAL ENGLISH FOR BPOS",
        " ENGLISH SKILLS":"ENGLISH SKILLS",
        "LOGISTICA Y CADENA DE ABASTECIMIENTO ":"LOGISTICA Y CADENA DE ABASTECIMIENTO",
        "LOGÍSTICA Y CADENA DE ABASTECIMIENTO":"LOGISTICA Y CADENA DE ABASTECIMIENTO",
        "REDACCION Y ORTOGRAFIA TECNICAS Y HABILIDADES PARA LA COMUNICACION ESCRITA":"REDACCION Y ORTOGRAFIA TECNICAS Y HABILIDADES PARA COMUNICACION ESCRITA",
        "GERENCIA DE PROYECTOS CON METODOLOGIA PMI ": "GERENCIA DE PROYECTOS CON METODOLOGIA PMI"
    }
    df_unificado['TALLER Y/O CURSO '] = df_unificado['TALLER Y/O CURSO '].replace(reemplazos)
    columnas_requeridas = ['SEDE PROVEEDOR', 'TIPO DE IDENTIFICACIÓN', 'NUMERO DE IDENTIFICACIÓN', 'GENERO', 
                       'PRIMER NOMBRE', 'SEGUNDO NOMBRE', 'PRIMER APELLIDO', 'SEGUNDO APELLIDO ', 
                       'CELULAR', 'TELEFONO', 'CORREO ELECTRÓNICO']
    df_unificado = df_unificado.dropna(subset=columnas_requeridas, how='all')
    return df_unificado

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def insertar_tablas_dinamicas(ruta_archivo):
    """
    Crea una hoja con los datos resumidos que servirán como base para crear tablas dinámicas en Excel.

    Args:
        ruta_archivo (str): Ruta del archivo Excel en el que se añadirán los datos base.
    """
    try:
        # Cargar el archivo Excel
        wb = load_workbook(ruta_archivo)

        # Crear o limpiar la hoja "TD"
        if "TD" in wb.sheetnames:
            ws_td = wb["TD"]
            wb.remove(ws_td)
        ws_td = wb.create_sheet("TD")

        # Cargar datos de la primera hoja
        ws_data = wb[wb.sheetnames[0]]  # Supone que los datos están en la primera hoja

        # Crear resumen para la tabla dinámica
        ws_td.append(["TALLER Y/O CURSO", "TIPO", "Recuento de IDENTIFICACIÓN", "Fecha Apertura Min", "Fecha Cierre Max"])
        resumen = {}

        for row in ws_data.iter_rows(min_row=2, values_only=True):  # Saltar encabezado
            curso = row[0]  # Suponiendo que "TALLER Y/O CURSO" está en la primera columna
            tipo = row[1]  # Suponiendo que "TIPO" está en la segunda columna
            identificacion = row[2]  # Suponiendo que "NUMERO DE IDENTIFICACIÓN" está en la tercera columna
            fecha_apertura = row[3]  # Suponiendo que "FECHA DE APERTURA" está en la cuarta columna
            fecha_cierre = row[4]  # Suponiendo que "FECHA DE CIERRE" está en la quinta columna

            if (curso, tipo) not in resumen:
                resumen[(curso, tipo)] = {
                    "count": 0,
                    "fecha_apertura_min": fecha_apertura,
                    "fecha_cierre_max": fecha_cierre,
                }

            resumen[(curso, tipo)]["count"] += 1
            if fecha_apertura and (resumen[(curso, tipo)]["fecha_apertura_min"] is None or fecha_apertura < resumen[(curso, tipo)]["fecha_apertura_min"]):
                resumen[(curso, tipo)]["fecha_apertura_min"] = fecha_apertura
            if fecha_cierre and (resumen[(curso, tipo)]["fecha_cierre_max"] is None or fecha_cierre > resumen[(curso, tipo)]["fecha_cierre_max"]):
                resumen[(curso, tipo)]["fecha_cierre_max"] = fecha_cierre

        # Escribir resumen en la hoja "TD"
        for (curso, tipo), values in resumen.items():
            ws_td.append([
                curso, 
                tipo, 
                values["count"], 
                values["fecha_apertura_min"], 
                values["fecha_cierre_max"]
            ])

        # Ajustar ancho de columnas
        for col in ws_td.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws_td.column_dimensions[col_letter].width = max_length + 2

        # Guardar los cambios
        wb.save(ruta_archivo)
        print("Tablas dinámicas base creadas exitosamente.")

    except Exception as e:
        print(f"Error al crear tablas dinámicas base: {e}")



# Interfaz Streamlit
st.title("ETL de Consolidación de Ciclos")

# Encabezado e instrucciones
st.header("""
          Fundación Universitaria Compensar
          Proyectos Especiales, Coordinación Operativa""")

# Sección desplegable con información
with st.expander("Acerca de esta herramienta"):
    st.markdown("""
    ### ¿Qué hace esta herramienta?
    - Extrae y procesa automáticamente archivos adjuntos de correos en formato `.eml`.
    - Consolida múltiples archivos de Excel en un único archivo organizado.
    - Genera tablas dinámicas en el archivo consolidado.
    - Comprime los archivos individuales en un archivo ZIP para su descarga.

    ### Instrucciones:
    1. **Sube tus archivos:** Arrastra y suelta tus correos en formato `.eml` con los Excel adjuntos; o un archivo comprimido con todos los archivos de Excel a consolidar.
    2. **Número de ciclo:** Ingresa el número que identificará el archivo consolidado.
    3. **Procesa:** Haz clic en **"Procesar archivos"** para ejecutar el flujo.
    4. **Descarga:** 
        - El archivo consolidado con tablas dinámicas.
        - Un archivo ZIP con los archivos individuales.

    ### Al utilizar esta herramienta recuerda:
    - Asegúrate de que tus correos `.eml` contengan archivos de Excel como adjuntos.
    - El archivo comprimido debe tener únicamente los archivos de Excel remitidos por la Agencia de Empleo.
    - Los resultados estarán disponibles hasta que cierres la aplicación.
    """)

# Subir archivos
st.header("Subir correos .eml o archivos comprimidos")
uploaded_files = st.file_uploader("Sube tus correos .eml o un archivo comprimido:", accept_multiple_files=True)

# Carpeta temporal para procesar datos
temp_dir = Path("temp")
temp_dir.mkdir(exist_ok=True)

# Entrada de usuario para el número de ciclo
numero_ciclo = st.text_input("Número de ciclo para guardar el archivo consolidado:", value="---")

# Carpeta temporal para procesar datos
temp_dir = Path("temp")
temp_dir.mkdir(exist_ok=True)

# Función para comprimir archivos individuales en un ZIP
def crear_zip(carpeta, archivo_zip):
    """
    Comprime todos los archivos .xlsx y .xls en la carpeta especificada
    dentro de un archivo ZIP.
    
    Args:
        carpeta (str or Path): Ruta de la carpeta donde están los archivos.
        archivo_zip (str or Path): Ruta donde se guardará el archivo ZIP.
    """
    import zipfile
    with zipfile.ZipFile(archivo_zip, 'w') as zf:
        for archivo in os.listdir(carpeta):
            if archivo.endswith('.xlsx') or archivo.endswith('.xls'):
                zf.write(os.path.join(carpeta, archivo), archivo)



# Procesar datos
if st.button("Procesar archivos"):
    if uploaded_files:
        try:
            # Crear carpeta para guardar los archivos extraídos
            carpeta_bases = temp_dir / "Bases"
            carpeta_bases.mkdir(parents=True, exist_ok=True)
            guardar_adjuntos_eml(uploaded_files, carpeta_bases)
            st.success("Archivos adjuntos extraídos correctamente.")

            # Consolidar archivos en un único DataFrame
            df_consolidado = consolidar_archivos(carpeta_bases)
            archivo_salida = temp_dir / f"Consolidado_Ciclo_{numero_ciclo}.xlsx"
            df_consolidado.to_excel(archivo_salida, index=False)
            st.success("Archivo consolidado generado correctamente.")

            # Insertar tablas dinámicas en el archivo consolidado
            insertar_tablas_dinamicas(archivo_salida)
            st.success("Tablas dinámicas insertadas correctamente.")

            # Crear un archivo ZIP con los Excel individuales
            archivo_zip = temp_dir / f"Archivos_Individuales_Ciclo_{numero_ciclo}.zip"
            crear_zip(carpeta_bases, archivo_zip)
            st.success("Archivo comprimido generado correctamente.")

            # Guardar rutas en `st.session_state`
            st.session_state['archivo_salida'] = archivo_salida
            st.session_state['archivo_zip'] = archivo_zip

        except Exception as e:
            st.error(f"Ocurrió un error al procesar los archivos: {e}")
    else:
        st.warning("Por favor, sube al menos un archivo para procesar.")

# Opciones de descarga
if 'archivo_salida' in st.session_state and 'archivo_zip' in st.session_state:
    with open(st.session_state['archivo_salida'], "rb") as f:
        st.download_button(
            label="Descargar archivo consolidado",
            data=f,
            file_name=f"Consolidado_Ciclo_{numero_ciclo}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    with open(st.session_state['archivo_zip'], "rb") as f:
        st.download_button(
            label="Descargar archivos comprimidos",
            data=f,
            file_name=f"Archivos_Individuales_Ciclo_{numero_ciclo}.zip",
            mime="application/zip",
        )
